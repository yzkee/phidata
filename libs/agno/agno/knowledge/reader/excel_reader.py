import asyncio
import io
from pathlib import Path
from typing import IO, Any, Iterable, List, Optional, Sequence, Tuple, Union

from agno.knowledge.chunking.row import RowChunking
from agno.knowledge.chunking.strategy import ChunkingStrategy, ChunkingStrategyType
from agno.knowledge.document.base import Document
from agno.knowledge.reader.base import Reader
from agno.knowledge.reader.utils import (
    convert_xls_cell_value,
    excel_rows_to_documents,
    get_workbook_name,
    infer_file_extension,
)
from agno.knowledge.types import ContentType
from agno.utils.log import log_debug, log_error


class ExcelReader(Reader):
    """Reader for Excel files (.xlsx and .xls)."""

    def __init__(
        self,
        sheets: Optional[List[Union[str, int]]] = None,
        chunking_strategy: Optional[ChunkingStrategy] = RowChunking(),
        **kwargs,
    ):
        super().__init__(chunking_strategy=chunking_strategy, **kwargs)
        self.sheets = sheets

    @classmethod
    def get_supported_chunking_strategies(cls) -> List[ChunkingStrategyType]:
        """Get the list of supported chunking strategies for Excel readers."""
        return [
            ChunkingStrategyType.ROW_CHUNKER,
            ChunkingStrategyType.CODE_CHUNKER,
            ChunkingStrategyType.FIXED_SIZE_CHUNKER,
            ChunkingStrategyType.AGENTIC_CHUNKER,
            ChunkingStrategyType.DOCUMENT_CHUNKER,
            ChunkingStrategyType.RECURSIVE_CHUNKER,
        ]

    @classmethod
    def get_supported_content_types(cls) -> List[ContentType]:
        """Get the list of supported content types."""
        return [ContentType.XLSX, ContentType.XLS]

    def _should_include_sheet(
        self,
        sheet_name: str,
        sheet_index: int,
    ) -> bool:
        """Check if sheet passes the configured filters.

        Args:
            sheet_name: Name of the sheet
            sheet_index: 1-based index of the sheet (matches document metadata)

        Returns:
            True if sheet should be included, False otherwise.

        Note:
            - Index filtering is 1-based to match sheet_index in document metadata
            - Name filtering is case-insensitive
            - Empty list or None means include all sheets
        """
        # None or empty list = include all sheets
        if not self.sheets:
            return True

        for sheet_filter in self.sheets:
            if isinstance(sheet_filter, int):
                # 1-based indexing to match metadata
                if sheet_index == sheet_filter:
                    return True
            elif isinstance(sheet_filter, str):
                # Case-insensitive name matching
                if sheet_name.lower() == sheet_filter.lower():
                    return True

        return False

    def _read_xlsx(self, file: Union[Path, IO[Any]], *, workbook_name: str) -> List[Document]:
        """Read .xlsx file using openpyxl."""
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError("`openpyxl` not installed. Please install it via `pip install openpyxl`.") from e

        if isinstance(file, Path):
            workbook = openpyxl.load_workbook(filename=str(file), read_only=True, data_only=True)
        else:
            file.seek(0)
            raw = file.read()
            if isinstance(raw, str):
                raw = raw.encode("utf-8", errors="replace")
            workbook = openpyxl.load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)

        try:
            sheets: List[Tuple[str, int, Iterable[Sequence[Any]]]] = []
            for sheet_index, worksheet in enumerate(workbook.worksheets):
                # Pass 1-based index to match metadata (sheet_index + 1)
                if not self._should_include_sheet(worksheet.title, sheet_index + 1):
                    log_debug(f"Skipping sheet '{worksheet.title}' (filtered out)")
                    continue

                sheets.append((worksheet.title, sheet_index + 1, worksheet.iter_rows(values_only=True)))

            return excel_rows_to_documents(workbook_name=workbook_name, sheets=sheets)
        finally:
            workbook.close()

    def _read_xls(self, file: Union[Path, IO[Any]], *, workbook_name: str) -> List[Document]:
        """Read .xls file using xlrd."""
        try:
            import xlrd
        except ImportError as e:
            raise ImportError("`xlrd` not installed. Please install it via `pip install xlrd`.") from e

        if isinstance(file, Path):
            workbook = xlrd.open_workbook(filename=str(file), encoding_override=self.encoding)
        else:
            file.seek(0)
            raw = file.read()
            if isinstance(raw, str):
                raw = raw.encode("utf-8", errors="replace")
            workbook = xlrd.open_workbook(file_contents=raw, encoding_override=self.encoding)

        sheets: List[Tuple[str, int, Iterable[Sequence[Any]]]] = []
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)

            # Pass 1-based index to match metadata (sheet_index + 1)
            if not self._should_include_sheet(sheet.name, sheet_index + 1):
                log_debug(f"Skipping sheet '{sheet.name}' (filtered out)")
                continue

            def _iter_sheet_rows(_sheet: Any = sheet, _datemode: int = workbook.datemode) -> Iterable[Sequence[Any]]:
                for row_index in range(_sheet.nrows):
                    yield [
                        convert_xls_cell_value(
                            _sheet.cell_value(row_index, col_index),
                            _sheet.cell_type(row_index, col_index),
                            _datemode,
                        )
                        for col_index in range(_sheet.ncols)
                    ]

            sheets.append((sheet.name, sheet_index + 1, _iter_sheet_rows()))

        return excel_rows_to_documents(workbook_name=workbook_name, sheets=sheets)

    def read(
        self,
        file: Union[Path, IO[Any]],
        name: Optional[str] = None,
    ) -> List[Document]:
        """Read an Excel file and return documents (one per sheet)."""
        try:
            file_extension = infer_file_extension(file, name)
            workbook_name = get_workbook_name(file, name)

            if isinstance(file, Path) and not file.exists():
                raise FileNotFoundError(f"Could not find file: {file}")

            file_desc = str(file) if isinstance(file, Path) else getattr(file, "name", "BytesIO")
            log_debug(f"Reading Excel file: {file_desc}")

            if file_extension == ContentType.XLSX or file_extension == ".xlsx":
                documents = self._read_xlsx(file, workbook_name=workbook_name)
            elif file_extension == ContentType.XLS or file_extension == ".xls":
                documents = self._read_xls(file, workbook_name=workbook_name)
            else:
                raise ValueError(f"Unsupported file extension: '{file_extension}'. Expected .xlsx or .xls")

            if self.chunk:
                chunked_documents = []
                for document in documents:
                    chunked_documents.extend(self.chunk_document(document))
                return chunked_documents

            return documents

        except (FileNotFoundError, ImportError, ValueError):
            raise
        except Exception as e:
            file_desc = getattr(file, "name", str(file)) if isinstance(file, IO) else file
            log_error(f"Error reading {file_desc}: {e}")
            return []

    async def async_read(
        self,
        file: Union[Path, IO[Any]],
        name: Optional[str] = None,
    ) -> List[Document]:
        """Async version of read()."""
        try:
            file_extension = infer_file_extension(file, name)
            workbook_name = get_workbook_name(file, name)

            if isinstance(file, Path) and not file.exists():
                raise FileNotFoundError(f"Could not find file: {file}")

            file_desc = str(file) if isinstance(file, Path) else getattr(file, "name", "BytesIO")
            log_debug(f"Reading Excel file async: {file_desc}")

            if file_extension == ContentType.XLSX or file_extension == ".xlsx":
                documents = await asyncio.to_thread(self._read_xlsx, file, workbook_name=workbook_name)
            elif file_extension == ContentType.XLS or file_extension == ".xls":
                documents = await asyncio.to_thread(self._read_xls, file, workbook_name=workbook_name)
            else:
                raise ValueError(f"Unsupported file extension: '{file_extension}'. Expected .xlsx or .xls")

            if self.chunk:
                documents = await self.chunk_documents_async(documents)

            return documents

        except (FileNotFoundError, ImportError, ValueError):
            raise
        except Exception as e:
            file_desc = getattr(file, "name", str(file)) if isinstance(file, IO) else file
            log_error(f"Error reading {file_desc}: {e}")
            return []
