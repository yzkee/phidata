import asyncio
import csv
import io
from datetime import date, datetime
from pathlib import Path
from typing import IO, Any, Iterable, List, Optional, Sequence, Tuple, Union
from uuid import uuid4

try:
    import aiofiles
except ImportError:
    raise ImportError("`aiofiles` not installed. Please install it with `pip install aiofiles`")

from agno.knowledge.chunking.row import RowChunking
from agno.knowledge.chunking.strategy import ChunkingStrategy, ChunkingStrategyType
from agno.knowledge.document.base import Document
from agno.knowledge.reader.base import Reader
from agno.knowledge.types import ContentType
from agno.utils.log import log_debug, log_error


def _get_workbook_name(file: Union[Path, IO[Any]], name: Optional[str]) -> str:
    """Extract workbook name from file path or name parameter.

    Priority: explicit name > file path stem > file object name attribute > "workbook"
    """
    if name:
        return Path(name).stem
    if isinstance(file, Path):
        return file.stem
    return Path(getattr(file, "name", "workbook")).stem


def _infer_file_extension(file: Union[Path, IO[Any]], name: Optional[str]) -> str:
    if isinstance(file, Path):
        return file.suffix.lower()

    file_name = getattr(file, "name", None)
    if isinstance(file_name, str) and file_name:
        return Path(file_name).suffix.lower()

    if name:
        return Path(name).suffix.lower()

    return ""


def _convert_xls_cell_value(cell_value: Any, cell_type: int, datemode: int) -> Any:
    """Convert xlrd cell value to Python type.

    xlrd returns dates as Excel serial numbers and booleans as 0/1 integers.
    This converts them to proper Python types for consistent handling with openpyxl.
    """
    try:
        import xlrd
    except ImportError:
        return cell_value

    if cell_type == xlrd.XL_CELL_DATE:
        try:
            date_tuple = xlrd.xldate_as_tuple(cell_value, datemode)
            return datetime(*date_tuple)
        except Exception:
            return cell_value
    if cell_type == xlrd.XL_CELL_BOOLEAN:
        return bool(cell_value)
    return cell_value


def _stringify_spreadsheet_cell_value(value: Any) -> str:
    if value is None:
        return ""

    # Handle datetime/date before float check (datetime is not a float)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    result = str(value)
    # Normalize all line endings to space to preserve row integrity in CSV-like output
    # Must handle CRLF first before individual CR/LF to avoid double-spacing
    result = result.replace("\r\n", " ")  # Windows (CRLF)
    result = result.replace("\r", " ")  # Old Mac (CR)
    result = result.replace("\n", " ")  # Unix (LF)
    return result


def _row_values_to_csv_line(row_values: Sequence[Any]) -> str:
    values = [_stringify_spreadsheet_cell_value(v) for v in row_values]
    while values and values[-1] == "":
        values.pop()

    return ", ".join(values)


def _excel_rows_to_documents(
    *,
    workbook_name: str,
    sheets: Iterable[Tuple[str, Iterable[Sequence[Any]]]],
) -> List[Document]:
    documents = []
    for sheet_index, (sheet_name, rows) in enumerate(sheets, start=1):
        lines = []
        for row in rows:
            line = _row_values_to_csv_line(row)
            if line:
                lines.append(line)

        if not lines:
            log_debug(f"Sheet '{sheet_name}' is empty, skipping")
            continue

        documents.append(
            Document(
                name=workbook_name,
                id=str(uuid4()),
                meta_data={"sheet_name": sheet_name, "sheet_index": sheet_index},
                content="\n".join(lines),
            )
        )

    return documents


class CSVReader(Reader):
    """Reader for CSV files"""

    def __init__(self, chunking_strategy: Optional[ChunkingStrategy] = RowChunking(), **kwargs):
        super().__init__(chunking_strategy=chunking_strategy, **kwargs)

    @classmethod
    def get_supported_chunking_strategies(cls) -> List[ChunkingStrategyType]:
        """Get the list of supported chunking strategies for CSV readers."""
        return [
            ChunkingStrategyType.ROW_CHUNKER,
            ChunkingStrategyType.CODE_CHUNKER,
            ChunkingStrategyType.FIXED_SIZE_CHUNKER,
            ChunkingStrategyType.AGENTIC_CHUNKER,
            ChunkingStrategyType.DOCUMENT_CHUNKER,
            ChunkingStrategyType.RECURSIVE_CHUNKER,
        ]

    @classmethod
    def get_supported_content_types(cls) -> List[ContentType]:
        return [ContentType.CSV, ContentType.XLSX, ContentType.XLS]

    def read(
        self, file: Union[Path, IO[Any]], delimiter: str = ",", quotechar: str = '"', name: Optional[str] = None
    ) -> List[Document]:
        try:
            file_extension = _infer_file_extension(file, name)
            if file_extension in {ContentType.XLSX, ContentType.XLS}:
                workbook_name = _get_workbook_name(file, name)

                if file_extension == ContentType.XLSX:
                    documents = self._read_xlsx(file, workbook_name=workbook_name)
                else:
                    documents = self._read_xls(file, workbook_name=workbook_name)

                if self.chunk:
                    chunked_documents = []
                    for document in documents:
                        chunked_documents.extend(self.chunk_document(document))
                    return chunked_documents
                return documents

            if isinstance(file, Path):
                if not file.exists():
                    raise FileNotFoundError(f"Could not find file: {file}")
                log_debug(f"Reading: {file}")
                csv_name = name or file.stem
                file_content: Union[io.TextIOWrapper, io.StringIO] = file.open(
                    newline="", mode="r", encoding=self.encoding or "utf-8"
                )
            else:
                log_debug(f"Reading retrieved file: {getattr(file, 'name', 'BytesIO')}")
                csv_name = name or getattr(file, "name", "csv_file").split(".")[0]
                file.seek(0)
                file_content = io.StringIO(file.read().decode(self.encoding or "utf-8"))

            csv_lines: List[str] = []
            with file_content as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=delimiter, quotechar=quotechar)
                for row in csv_reader:
                    # Use stringify to normalize line endings in CSV cells
                    csv_lines.append(", ".join(_stringify_spreadsheet_cell_value(cell) for cell in row))

            documents = [
                Document(
                    name=csv_name,
                    id=str(uuid4()),
                    content="\n".join(csv_lines),
                )
            ]
            if self.chunk:
                chunked_documents = []
                for document in documents:
                    chunked_documents.extend(self.chunk_document(document))
                return chunked_documents
            return documents
        except FileNotFoundError:
            raise
        except ImportError:
            raise
        except UnicodeDecodeError as e:
            file_desc = getattr(file, "name", str(file)) if isinstance(file, IO) else file
            log_error(f"Encoding error reading {file_desc}: {e}. Try specifying a different encoding.")
            return []
        except Exception as e:
            file_desc = getattr(file, "name", str(file)) if isinstance(file, IO) else file
            log_error(f"Error reading {file_desc}: {e}")
            return []

    async def async_read(
        self,
        file: Union[Path, IO[Any]],
        delimiter: str = ",",
        quotechar: str = '"',
        page_size: int = 1000,
        name: Optional[str] = None,
    ) -> List[Document]:
        """
        Read a CSV file asynchronously, processing batches of rows concurrently.

        Args:
            file: Path or file-like object
            delimiter: CSV delimiter
            quotechar: CSV quote character
            page_size: Number of rows per page

        Returns:
            List of Document objects
        """
        try:
            file_extension = _infer_file_extension(file, name)
            if file_extension in {ContentType.XLSX, ContentType.XLS}:
                workbook_name = _get_workbook_name(file, name)

                if file_extension == ContentType.XLSX:
                    documents = await asyncio.to_thread(self._read_xlsx, file, workbook_name=workbook_name)
                else:
                    documents = await asyncio.to_thread(self._read_xls, file, workbook_name=workbook_name)

                if self.chunk:
                    documents = await self.chunk_documents_async(documents)
                return documents

            if isinstance(file, Path):
                if not file.exists():
                    raise FileNotFoundError(f"Could not find file: {file}")
                log_debug(f"Reading async: {file}")
                async with aiofiles.open(file, mode="r", encoding=self.encoding or "utf-8", newline="") as file_content:
                    content = await file_content.read()
                    file_content_io = io.StringIO(content)
                csv_name = name or file.stem
            else:
                log_debug(f"Reading retrieved file async: {getattr(file, 'name', 'BytesIO')}")
                file.seek(0)
                file_content_io = io.StringIO(file.read().decode(self.encoding or "utf-8"))
                csv_name = name or getattr(file, "name", "csv_file").split(".")[0]

            file_content_io.seek(0)
            csv_reader = csv.reader(file_content_io, delimiter=delimiter, quotechar=quotechar)
            rows = list(csv_reader)
            total_rows = len(rows)

            if total_rows <= 10:
                # Use stringify to normalize line endings in CSV cells
                csv_content = " ".join(
                    ", ".join(_stringify_spreadsheet_cell_value(cell) for cell in row) for row in rows
                )
                documents = [
                    Document(
                        name=csv_name,
                        id=str(uuid4()),
                        content=csv_content,
                    )
                ]
            else:
                pages = []
                for i in range(0, total_rows, page_size):
                    pages.append(rows[i : i + page_size])

                async def _process_page(page_number: int, page_rows: List[List[str]]) -> Document:
                    """Process a page of rows into a document"""
                    start_row = (page_number - 1) * page_size + 1
                    # Use stringify to normalize line endings in CSV cells
                    page_content = " ".join(
                        ", ".join(_stringify_spreadsheet_cell_value(cell) for cell in row) for row in page_rows
                    )

                    return Document(
                        name=csv_name,
                        id=str(uuid4()),
                        meta_data={"page": page_number, "start_row": start_row, "rows": len(page_rows)},
                        content=page_content,
                    )

                documents = await asyncio.gather(
                    *[_process_page(page_number, page) for page_number, page in enumerate(pages, start=1)]
                )

            if self.chunk:
                documents = await self.chunk_documents_async(documents)

            return documents
        except FileNotFoundError:
            raise
        except ImportError:
            raise
        except UnicodeDecodeError as e:
            file_desc = getattr(file, "name", str(file)) if isinstance(file, IO) else file
            log_error(f"Encoding error reading {file_desc}: {e}. Try specifying a different encoding.")
            return []
        except Exception as e:
            file_desc = getattr(file, "name", str(file)) if isinstance(file, IO) else file
            log_error(f"Error reading {file_desc}: {e}")
            return []

    def _read_xlsx(self, file: Union[Path, IO[Any]], *, workbook_name: str) -> List[Document]:
        try:
            import openpyxl  # type: ignore
        except ImportError as e:
            raise ImportError(
                "`openpyxl` not installed. Please install it via `pip install agno[csv]` or `pip install openpyxl`."
            ) from e

        if isinstance(file, Path):
            workbook = openpyxl.load_workbook(filename=str(file), read_only=True, data_only=True)
        else:
            file.seek(0)
            raw = file.read()
            if isinstance(raw, str):
                raw = raw.encode("utf-8", errors="replace")
            workbook = openpyxl.load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)

        try:
            return _excel_rows_to_documents(
                workbook_name=workbook_name,
                sheets=[(worksheet.title, worksheet.iter_rows(values_only=True)) for worksheet in workbook.worksheets],
            )
        finally:
            workbook.close()

    def _read_xls(self, file: Union[Path, IO[Any]], *, workbook_name: str) -> List[Document]:
        try:
            import xlrd  # type: ignore
        except ImportError as e:
            raise ImportError(
                "`xlrd` not installed. Please install it via `pip install agno[csv]` or `pip install xlrd`."
            ) from e

        if isinstance(file, Path):
            workbook = xlrd.open_workbook(filename=str(file))
        else:
            file.seek(0)
            raw = file.read()
            if isinstance(raw, str):
                raw = raw.encode("utf-8", errors="replace")
            workbook = xlrd.open_workbook(file_contents=raw)

        sheets: List[Tuple[str, Iterable[Sequence[Any]]]] = []
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)

            def _iter_sheet_rows(_sheet: Any = sheet, _datemode: int = workbook.datemode) -> Iterable[Sequence[Any]]:
                for row_index in range(_sheet.nrows):
                    yield [
                        _convert_xls_cell_value(
                            _sheet.cell_value(row_index, col_index),
                            _sheet.cell_type(row_index, col_index),
                            _datemode,
                        )
                        for col_index in range(_sheet.ncols)
                    ]

            sheets.append((sheet.name, _iter_sheet_rows()))

        return _excel_rows_to_documents(workbook_name=workbook_name, sheets=sheets)
